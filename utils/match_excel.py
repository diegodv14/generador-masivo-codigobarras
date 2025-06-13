from openpyxl import load_workbook

def match_excel_existing(codes: list[str]):

    wb = load_workbook("GAF-RGT-008_INVENTARIO_EQUIPOS TECNOLOGICO_MUEBLES.xlsx")

    ws_mue = wb["MUEBLES DE OFIC"]
    ws_hwc = wb["EQUIPOS COMPUTO"]
    ws_ct = wb["BASES CELULARES"]

    mue_codes = [c for c in codes if c.startswith("MUE")]
    hwc_codes = [c for c in codes if c.startswith("ECT")]
    ct_codes = [c for c in codes if c.startswith("BCT")] 

    start_row_mue = 15
    col_mue = 5
    for i, code in enumerate(mue_codes):
        ws_mue.cell(row=start_row_mue + i, column=col_mue, value=code)
    start_row_hwc = 15
    col_hwc = 5
    for i, code in enumerate(hwc_codes):
        ws_hwc.cell(row=start_row_hwc + i, column=col_hwc, value=code)
        
    start_row_ct = 14
    col_ct = 7
    for i, code in enumerate(ct_codes):
        ws_ct.cell(row=start_row_ct + i, column=col_ct, value=code)
    wb.save("GAF-RGT-008_INVENTARIO_EQUIPOS TECNOLOGICO_MUEBLES.xlsx")
    print("Excel actualizado con los codigos de barras correctamente")
